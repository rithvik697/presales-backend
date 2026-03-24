import csv
import io
from typing import Any, Dict, List, Tuple

from db import get_db
from services.leads_service import add_new_lead
from services.webhook_service import (
    _auto_assign_employee,
    _find_project_by_name,
    _find_source_by_name,
    _update_assignment_tracker,
)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency for xlsx support
    load_workbook = None


EXPECTED_COLUMNS = [
    "first_name",
    "last_name",
    "phone",
    "email",
    "project_name",
    "source_name",
    "description",
    "alternate_phone",
    "profession",
    "assigned_to",
    "username",
    "employee_name",
]

LEGACY_HEADERS = {
    "customer_name",
    "country_code",
    "mobile_number",
    "requirement_name",
    "lead_source",
}


def _ensure_bulk_upload_log_table(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lead_bulk_upload_log (
            upload_id INT AUTO_INCREMENT PRIMARY KEY,
            file_name VARCHAR(255) NOT NULL,
            total_rows INT NOT NULL DEFAULT 0,
            created_count INT NOT NULL DEFAULT 0,
            duplicate_count INT NOT NULL DEFAULT 0,
            failed_count INT NOT NULL DEFAULT 0,
            uploaded_by VARCHAR(150) NOT NULL,
            uploaded_on DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _get_row_value(row: Dict[str, Any], *keys: str) -> str:
    for key in keys:
        if key in row and row[key] not in [None, ""]:
            return str(row[key]).strip()
    return ""


def _read_csv_rows(file_storage) -> List[Dict[str, str]]:
    content = file_storage.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        return []

    rows = []
    for raw_row in reader:
        row = {}
        for key, value in raw_row.items():
            row[_normalize_header(key)] = str(value).strip() if value is not None else ""
        rows.append(row)
    return rows


def _read_xlsx_rows(file_storage) -> List[Dict[str, str]]:
    if load_workbook is None:
        raise ValueError("XLSX upload requires openpyxl to be installed on the backend")

    workbook = load_workbook(filename=io.BytesIO(file_storage.read()), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(cell) for cell in rows[0]]
    data_rows: List[Dict[str, str]] = []

    for row_values in rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row_values[index] if index < len(row_values) else ""
            row[header] = str(value).strip() if value is not None else ""
        data_rows.append(row)

    return data_rows


def _parse_rows(file_storage) -> List[Dict[str, str]]:
    filename = (file_storage.filename or "").lower()
    file_storage.stream.seek(0)

    if filename.endswith(".csv"):
        return _read_csv_rows(file_storage)

    if filename.endswith(".xlsx"):
        return _read_xlsx_rows(file_storage)

    raise ValueError("Only .csv and .xlsx files are supported")


def _is_legacy_row_format(rows: List[Dict[str, str]]) -> bool:
    if not rows:
        return False
    headers = set(rows[0].keys())
    return LEGACY_HEADERS.issubset(headers)


def _prefix_country_code(country_code: str) -> str:
    digits = "".join(ch for ch in str(country_code or "").strip() if ch.isdigit())
    return f"+{digits}" if digits else ""


def _legacy_phone_value(row: Dict[str, str]) -> str:
    country_code = _prefix_country_code(_get_row_value(row, "country_code"))
    phone = _get_row_value(row, "mobile_number", "phone")
    return f"{country_code}{phone}" if country_code and phone else phone


def _legacy_alternate_phone_value(row: Dict[str, str]) -> str:
    alternate = _get_row_value(row, "alternate_number", "2nd_contact_num", "alternate_phone")
    if not alternate or alternate in {"0", "0000000000"}:
        return ""
    country_code = _prefix_country_code(_get_row_value(row, "country_code"))
    return f"{country_code}{alternate}" if country_code and alternate else alternate


def _map_legacy_row(row: Dict[str, str]) -> Dict[str, str]:
    full_name = _get_row_value(row, "customer_name", "name")
    full_name_parts = [part for part in full_name.split() if part]

    first_name = full_name_parts[0] if full_name_parts else ""
    derived_last_name = " ".join(full_name_parts[1:]) if len(full_name_parts) > 1 else ""
    explicit_last_name = _get_row_value(row, "last_name")

    description = _get_row_value(row, "last_remarks", "description")
    if description in {"0", "select"}:
        description = ""

    email = _get_row_value(row, "email_id", "email")
    if email == "0":
        email = ""

    return {
        "name": full_name,
        "first_name": first_name,
        "last_name": explicit_last_name or derived_last_name,
        "phone": _legacy_phone_value(row),
        "alternate_phone": _legacy_alternate_phone_value(row),
        "email": email,
        "project_name": _get_row_value(row, "requirement_name", "project_name"),
        "source_name": _get_row_value(row, "lead_source", "source_name"),
        "description": description,
        "profession": _get_row_value(row, "profession"),
        "employee_name": _get_row_value(row, "emp_name", "employee_name", "assigned_to"),
        "username": _get_row_value(row, "username"),
        "emp_id": _get_row_value(row, "emp_id"),
    }


def _prepare_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    if _is_legacy_row_format(rows):
        return [_map_legacy_row(row) for row in rows]
    return rows


def _get_new_enquiry_status(cursor) -> str:
    cursor.execute("""
        SELECT status_id
        FROM lead_status
        WHERE LOWER(status_name) = LOWER(%s)
          AND is_active = 1
        LIMIT 1
    """, ("New Enquiry",))
    result = cursor.fetchone()
    if not result:
        raise ValueError("Status 'New Enquiry' is not configured in the system")
    return result["status_id"]


def _match_required_ids(cursor, row: Dict[str, str]) -> Tuple[str, str]:
    project_name = _get_row_value(row, "project_name", "project")
    source_name = _get_row_value(row, "source_name", "source")

    if not project_name:
        raise ValueError("Project name is required")
    if not source_name:
        raise ValueError("Source name is required")

    project_id = _find_project_by_name(cursor, project_name)
    if not project_id:
        raise ValueError(f"Project '{project_name}' not found")

    source_id = _find_source_by_name(cursor, source_name)
    if not source_id:
        raise ValueError(f"Source '{source_name}' not found")

    return project_id, source_id


def _normalize_employee_token(value: str) -> str:
    token = str(value or "").strip()
    if not token:
        return ""
    if "-" in token:
        token = token.split("-", 1)[1].strip()
    return " ".join(token.lower().split())


def _resolve_explicit_assignee(cursor, row: Dict[str, str]) -> str:
    emp_id = _get_row_value(row, "emp_id", "assigned_to")
    username = _get_row_value(row, "username")
    employee_name = _get_row_value(row, "employee_name", "emp_name")

    if emp_id:
        cursor.execute("""
            SELECT emp_id
            FROM employee
            WHERE emp_id = %s
              AND role_id = 'SALES_EXEC'
              AND emp_status = 'Active'
            LIMIT 1
        """, (emp_id,))
        result = cursor.fetchone()
        if result:
            return result["emp_id"]
        raise ValueError(f"Assigned employee '{emp_id}' is not an active sales executive")

    if username:
        cursor.execute("""
            SELECT emp_id
            FROM employee
            WHERE LOWER(username) = LOWER(%s)
              AND role_id = 'SALES_EXEC'
              AND emp_status = 'Active'
            LIMIT 1
        """, (username,))
        result = cursor.fetchone()
        if result:
            return result["emp_id"]
        raise ValueError(f"Username '{username}' is not an active sales executive")

    if employee_name:
        normalized_name = _normalize_employee_token(employee_name)
        cursor.execute("""
            SELECT
                emp_id,
                LOWER(TRIM(CONCAT(emp_first_name, ' ', IFNULL(emp_last_name, '')))) AS full_name,
                LOWER(emp_first_name) AS first_name
            FROM employee
            WHERE role_id = 'SALES_EXEC'
              AND emp_status = 'Active'
        """)
        employees = cursor.fetchall()
        for employee in employees:
            full_name = " ".join((employee["full_name"] or "").split())
            first_name = " ".join((employee["first_name"] or "").split())
            if normalized_name in {full_name, first_name}:
                return employee["emp_id"]
        raise ValueError(f"Employee '{employee_name}' is not an active sales executive")

    return ""


def _build_lead_payload(cursor, row: Dict[str, str]) -> Tuple[Dict[str, str], bool]:
    first_name = _get_row_value(row, "first_name")
    last_name = _get_row_value(row, "last_name")
    name = _get_row_value(row, "name")

    if not name:
        name = " ".join(part for part in [first_name, last_name] if part).strip()

    if not name:
        raise ValueError("First name or full name is required")

    phone = _get_row_value(row, "phone", "mobile", "phone_number")
    if not phone:
        raise ValueError("Phone number is required")

    project_id, source_id = _match_required_ids(cursor, row)
    explicit_assignee = _resolve_explicit_assignee(cursor, row)
    assigned_to = explicit_assignee or _auto_assign_employee(cursor, project_id)
    status_id = _get_new_enquiry_status(cursor)

    return ({
        "name": name,
        "phone": phone,
        "email": _get_row_value(row, "email"),
        "project": project_id,
        "source": source_id,
        "status": status_id,
        "assigned_to": assigned_to,
        "description": _get_row_value(row, "description", "remarks"),
        "alternate_phone": _get_row_value(row, "alternate_phone", "alternate_number", "alt_num"),
        "profession": _get_row_value(row, "profession"),
    }, bool(explicit_assignee))


def _log_bulk_upload(cursor, file_name: str, total_rows: int, created_count: int, duplicate_count: int, failed_count: int, uploaded_by: str) -> int:
    _ensure_bulk_upload_log_table(cursor)
    cursor.execute("""
        INSERT INTO lead_bulk_upload_log
            (file_name, total_rows, created_count, duplicate_count, failed_count, uploaded_by)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (file_name, total_rows, created_count, duplicate_count, failed_count, uploaded_by))
    return cursor.lastrowid


def process_bulk_lead_upload(file_storage, actor_id: str) -> Dict[str, Any]:
    if not file_storage or not file_storage.filename:
        raise ValueError("Upload file is required")

    rows = _prepare_rows(_parse_rows(file_storage))
    if not rows:
        raise ValueError("The uploaded file is empty")

    conn = get_db()
    if not conn:
        raise Exception("DB connection failed")

    created_leads: List[Dict[str, str]] = []
    duplicate_rows: List[Dict[str, Any]] = []
    failed_rows: List[Dict[str, Any]] = []

    try:
        for index, row in enumerate(rows, start=2):
            cursor = conn.cursor(dictionary=True)
            try:
                lead_payload, used_explicit_assignee = _build_lead_payload(cursor, row)
                assigned_to = lead_payload["assigned_to"]
                project_id = lead_payload["project"]

                cursor.close()
                lead_id = add_new_lead(lead_payload, actor_id=actor_id, role="ADMIN")

                if not used_explicit_assignee:
                    tracker_cursor = conn.cursor(dictionary=True)
                    _update_assignment_tracker(tracker_cursor, project_id, assigned_to)
                    tracker_cursor.close()

                conn.commit()

                created_leads.append({
                    "row_number": index,
                    "lead_id": lead_id,
                    "assigned_to": assigned_to
                })
            except ValueError as exc:
                error_text = str(exc)
                if "already exists" in error_text.lower():
                    duplicate_rows.append({
                        "row_number": index,
                        "phone": _get_row_value(row, "phone", "mobile", "phone_number"),
                        "reason": error_text
                    })
                else:
                    failed_rows.append({
                        "row_number": index,
                        "phone": _get_row_value(row, "phone", "mobile", "phone_number"),
                        "reason": error_text
                    })
                conn.rollback()
            except Exception as exc:
                failed_rows.append({
                    "row_number": index,
                    "phone": _get_row_value(row, "phone", "mobile", "phone_number"),
                    "reason": str(exc)
                })
                conn.rollback()
            finally:
                try:
                    cursor.close()
                except Exception:
                    pass

        log_cursor = conn.cursor()
        upload_id = _log_bulk_upload(
            log_cursor,
            file_storage.filename,
            len(rows),
            len(created_leads),
            len(duplicate_rows),
            len(failed_rows),
            actor_id,
        )
        conn.commit()
        log_cursor.close()

        return {
            "upload_id": upload_id,
            "file_name": file_storage.filename,
            "total_rows": len(rows),
            "created_count": len(created_leads),
            "duplicate_count": len(duplicate_rows),
            "failed_count": len(failed_rows),
            "created_leads": created_leads,
            "duplicate_rows": duplicate_rows,
            "failed_rows": failed_rows,
            "expected_columns": EXPECTED_COLUMNS,
        }
    finally:
        conn.close()


def get_bulk_upload_history() -> List[Dict[str, Any]]:
    conn = get_db()
    if not conn:
        return []

    try:
        cursor = conn.cursor(dictionary=True)
        _ensure_bulk_upload_log_table(cursor)
        cursor.execute("""
            SELECT
                upload_id,
                file_name,
                total_rows,
                created_count,
                duplicate_count,
                failed_count,
                uploaded_by,
                uploaded_on
            FROM lead_bulk_upload_log
            ORDER BY uploaded_on DESC
        """)
        return cursor.fetchall()
    finally:
        conn.close()
